"""
Step 5: Excel export (premium redesign)
--------------------------------------------
Exports the full DCF analysis into a polished, multi-sheet Excel workbook:
Cover, Summary, AI Valuation Summary (optional), DCF Forecast, Sensitivity,
and Scenario Comparison.

Design note: this workbook contains computed VALUES (from Python), not live
Excel formulas - it's a generated report/output, not a fill-in template the
user edits and recalculates. Formatting conventions below (navy/gold accent
palette, professional font, consistent banners) are chosen for a polished
reading experience, not for the "blue=input, black=formula" convention used
in editable financial models.
"""

import math
from datetime import date
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# --- Palette: dark navy + gold accent, a "premium research report" feel ---
NAVY = "0B2545"
NAVY_LIGHT = "13315C"
GOLD = "C9A227"
LIGHT_FILL = "EEF2F7"
AMBER_FILL = "FFF6E3"
WHITE = "FFFFFF"
TEXT_DARK = "1F2937"

FONT_NAME = "Arial"

BANNER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
BANNER_TITLE_FONT = Font(name=FONT_NAME, color=WHITE, bold=True, size=20)
BANNER_SUBTITLE_FONT = Font(name=FONT_NAME, color=GOLD, bold=False, size=11)

HEADER_FILL = PatternFill(start_color=NAVY_LIGHT, end_color=NAVY_LIGHT, fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, color=WHITE, bold=True, size=11)
SECTION_FILL = PatternFill(start_color=LIGHT_FILL, end_color=LIGHT_FILL, fill_type="solid")
SECTION_FONT = Font(name=FONT_NAME, color=NAVY, bold=True, size=12)

TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color=NAVY)
LABEL_FONT = Font(name=FONT_NAME, bold=True, color=TEXT_DARK)
BODY_FONT = Font(name=FONT_NAME, color=TEXT_DARK, size=10.5)
DISCLAIMER_FONT = Font(name=FONT_NAME, italic=True, color=TEXT_DARK, size=9.5)

THIN_BORDER = Border(*[Side(style="thin", color="D1D5DB")] * 4)
GOLD_BORDER = Border(*[Side(style="thin", color=GOLD)] * 4)
AMBER_FILL_STYLE = PatternFill(start_color=AMBER_FILL, end_color=AMBER_FILL, fill_type="solid")

DISCLAIMER_TEXT = (
    "This workbook is a personal, educational portfolio project. All valuations, ratings, "
    "and AI-generated commentary are outputs of a simplified discounted cash flow model built "
    "on a specific set of assumptions - they are NOT financial advice and should not be relied "
    "upon to make any investment decision. The author accepts no responsibility for decisions "
    "made based on this content. Always consult a licensed financial advisor and conduct your "
    "own research before investing."
)


def _draw_banner(ws, title: str, subtitle: str, num_cols: int = 8, height: int = 42):
    """Full-width dark navy banner with a bold title and a gold subtitle line
    underneath - the visual signature used at the top of every sheet."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    for row in (1, 2):
        for col in range(1, num_cols + 1):
            ws.cell(row=row, column=col).fill = BANNER_FILL
    ws.cell(row=1, column=1, value=title).font = BANNER_TITLE_FONT
    ws.cell(row=1, column=1).alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.cell(row=2, column=1, value=subtitle).font = BANNER_SUBTITLE_FONT
    ws.cell(row=2, column=1).alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = height
    ws.row_dimensions[2].height = 20


def _draw_wrapped_box(ws, start_row: int, text: str, num_cols: int = 8,
                       fill=None, border=THIN_BORDER, font=BODY_FONT,
                       chars_per_line: int = 100) -> int:
    """Merge a block of cells, write wrapped text into it, and auto-size the
    row height based on estimated line count. Returns the row AFTER the box."""
    lines = max(2, math.ceil(len(text) / chars_per_line))
    height = lines * 15 + 10
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=num_cols)
    cell = ws.cell(row=start_row, column=1, value=text)
    cell.font = font
    cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left", indent=1)
    if fill:
        for col in range(1, num_cols + 1):
            ws.cell(row=start_row, column=col).fill = fill
    ws.row_dimensions[start_row].height = height
    return start_row + 1


def _style_header_row(ws, row_num: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _autofit_columns(ws, min_width=10, max_width=30):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, min_width), max_width)


# ---------------------------------------------------------------
# Cover sheet
# ---------------------------------------------------------------
def build_cover_sheet(ws, ticker: str, company_name: str):
    ws.sheet_view.showGridLines = False
    _draw_banner(ws, "AI-Assisted DCF Valuation Tool",
                 f"{company_name} ({ticker})  |  Generated {date.today().strftime('%d %B %Y')}",
                 height=54)

    r = 5
    ws.cell(row=r, column=1, value="About this workbook").font = SECTION_FONT
    for c in range(1, 9):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1
    r = _draw_wrapped_box(ws, r,
        "A Python-built pipeline that pulls live financial data, forecasts free cash flow, "
        "runs a discounted cash flow valuation with WACC/terminal-growth sensitivity analysis, "
        "and uses an AI layer to explain (never generate) the resulting numbers. Built as a "
        "personal portfolio project applying valuation methods from an investment analyst "
        "internship in a reproducible, code-based form.")
    r += 1

    ws.cell(row=r, column=1, value="Contents").font = SECTION_FONT
    for c in range(1, 9):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1
    for label in ["Summary", "Live DCF (Editable)", "Management Assumptions", "Analyst Insights",
                  "AI Valuation Summary", "Comps & Football Field", "Model Validation",
                  "DCF Forecast", "Sensitivity & Scenarios"]:
        ws.cell(row=r, column=1, value=f"\u2022  {label}").font = BODY_FONT
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Important Disclaimer").font = SECTION_FONT
    for c in range(1, 9):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1
    r = _draw_wrapped_box(ws, r, DISCLAIMER_TEXT, fill=AMBER_FILL_STYLE,
                           border=GOLD_BORDER, font=DISCLAIMER_FONT, chars_per_line=95)

    for col_letter, width in zip("ABCDEFGH", [4, 16, 16, 16, 16, 16, 16, 16]):
        ws.column_dimensions[col_letter].width = width


# ---------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------
def build_summary_sheet(ws, ticker, company_name, current_price, result, assumptions,
                          cash, total_debt, shares_outstanding, capm_info=None):
    _draw_banner(ws, f"{ticker} \u2014 Valuation Summary", company_name)

    upside = (result["implied_share_price"] / current_price - 1) * 100
    rows = [
        ("Current Price", current_price, "$#,##0.00"),
        ("Implied Share Price (DCF)", result["implied_share_price"], "$#,##0.00"),
        ("Upside / (Downside)", upside / 100, "0.0%"),
        ("", None, None),
        ("Enterprise Value", result["enterprise_value"], "$#,##0,,\" M\""),
        ("Equity Value", result["equity_value"], "$#,##0,,\" M\""),
        ("Cash", cash, "$#,##0,,\" M\""),
        ("Total Debt", total_debt, "$#,##0,,\" M\""),
        ("Shares Outstanding", shares_outstanding, "#,##0,,\" M\""),
        ("", None, None),
        ("Key Assumptions", None, None),
        ("Revenue Growth", assumptions["avg_revenue_growth"], "0.0%"),
        ("EBIT Margin", assumptions["avg_ebit_margin"], "0.0%"),
        ("Capex % of Revenue", assumptions["avg_capex_pct_revenue"], "0.0%"),
        ("D&A % of Revenue", assumptions["avg_da_pct_revenue"], "0.0%"),
        ("NWC % of Revenue", assumptions.get("avg_nwc_pct_revenue"), "0.0%"),
    ]

    r = 4
    for label, value, fmt in rows:
        is_section = label == "Key Assumptions"
        cell = ws.cell(row=r, column=1, value=label)
        cell.font = SECTION_FONT if is_section else LABEL_FONT
        if is_section:
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = SECTION_FILL
        if value is not None:
            vcell = ws.cell(row=r, column=2, value=value)
            if fmt:
                vcell.number_format = fmt
            vcell.font = BODY_FONT
        r += 1

    if capm_info:
        r += 1
        ws.cell(row=r, column=1, value="Suggested WACC (CAPM) \u2014 for reference").font = SECTION_FONT
        for c in range(1, 4):
            ws.cell(row=r, column=c).fill = SECTION_FILL
        r += 1
        ws.cell(row=r, column=1,
                value="This is a data-driven starting point, not an automatic override. "
                      "The WACC actually used in this valuation is shown above under Key Assumptions "
                      "and can be set independently.")
        ws.cell(row=r, column=1).font = DISCLAIMER_FONT
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 28
        r += 1

        capm_rows = [
            ("Beta" + (" (default: not available)" if capm_info["beta_was_missing"] else ""),
             capm_info["beta"], "0.00"),
            ("Risk-Free Rate (10Y Treasury)", capm_info["risk_free_rate"], "0.00%"),
            ("Equity Risk Premium", capm_info["equity_risk_premium"], "0.00%"),
            ("Cost of Equity", capm_info["cost_of_equity"], "0.00%"),
            ("Cost of Debt (after-tax)", capm_info["cost_of_debt_aftertax"], "0.00%"),
            ("Weight of Equity", capm_info["weight_equity"], "0.0%"),
            ("Weight of Debt", capm_info["weight_debt"], "0.0%"),
            ("Suggested WACC (CAPM)", capm_info["wacc"], "0.00%"),
        ]
        for label, value, fmt in capm_rows:
            is_result = label == "Suggested WACC (CAPM)"
            ws.cell(row=r, column=1, value=label).font = (
                Font(name=FONT_NAME, bold=True, color=NAVY) if is_result else LABEL_FONT
            )
            vcell = ws.cell(row=r, column=2, value=value)
            vcell.number_format = fmt
            vcell.font = Font(name=FONT_NAME, bold=True, color=NAVY) if is_result else BODY_FONT
            r += 1

    _autofit_columns(ws)


# ---------------------------------------------------------------
# AI Valuation Summary sheet
# ---------------------------------------------------------------
def build_ai_summary_sheet(ws, ticker, company_name, result, current_price,
                             gap_pct, rating, explanation, analyst_info=None):
    _draw_banner(ws, f"{ticker} \u2014 AI Valuation Summary", company_name)

    r = 4
    stars_display = "\u2605" * rating["stars"] + "\u2606" * (5 - rating["stars"])
    ws.cell(row=r, column=1, value=stars_display).font = Font(name=FONT_NAME, size=24, color=GOLD, bold=True)
    ws.cell(row=r, column=4, value=rating["label"]).font = Font(name=FONT_NAME, size=16, bold=True, color=NAVY)
    ws.row_dimensions[r].height = 32
    r += 1
    ws.cell(row=r, column=1,
            value="This rating reflects ONLY this model's own DCF, using conservative unadjusted "
                  "historical-median assumptions - it is a narrower claim than a general investment "
                  "recommendation. See the comparison below.").font = DISCLAIMER_FONT
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 28
    r += 2

    metrics = [
        ("Implied Share Price (DCF)", result["implied_share_price"], "$#,##0.00"),
        ("Current Market Price", current_price, "$#,##0.00"),
        ("Gap", gap_pct, "+0.0%;-0.0%"),
    ]
    for label, value, fmt in metrics:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        vcell = ws.cell(row=r, column=2, value=value)
        vcell.number_format = fmt
        vcell.font = BODY_FONT
        r += 1
    r += 1

    # Model View vs. Street View comparison - makes the divergence between
    # this model's narrow DCF-only claim and broader analyst consensus
    # explicit and intentional, rather than something the reader has to
    # notice and reconcile themselves across separate sheets.
    ws.cell(row=r, column=1, value="Model View vs. Street View").font = SECTION_FONT
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1

    headers = ["", "View", "Basis"]
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=r, column=col_num, value=header)
    _style_header_row(ws, r, 3)
    r += 1

    ws.cell(row=r, column=1, value="Model (this tool)").font = LABEL_FONT
    ws.cell(row=r, column=2, value=f"{rating['stars']}/5 \u2014 {rating['label']}").font = BODY_FONT
    ws.cell(row=r, column=3, value="Unadjusted historical-median DCF only").font = BODY_FONT
    for c in range(1, 4):
        ws.cell(row=r, column=c).border = THIN_BORDER
    r += 1

    street_view, street_basis = "Not available", "\u2014"
    if analyst_info and analyst_info.get("recommendations") is not None:
        try:
            latest = analyst_info["recommendations"].iloc[0]
            counts = {"Strong Buy": latest.get("strongBuy", 0), "Buy": latest.get("buy", 0),
                      "Hold": latest.get("hold", 0), "Sell": latest.get("sell", 0),
                      "Strong Sell": latest.get("strongSell", 0)}
            dominant = max(counts, key=counts.get)
            total = sum(counts.values())
            street_view = f"{dominant} ({counts[dominant]}/{total} analysts)"
            mean_target = analyst_info.get("price_targets", {}).get("mean")
            street_basis = ("Analyst consensus" + (f", mean target ${mean_target:,.2f}" if mean_target else ""))
        except Exception:
            pass

    ws.cell(row=r, column=1, value="Street (analyst consensus)").font = LABEL_FONT
    ws.cell(row=r, column=2, value=street_view).font = BODY_FONT
    ws.cell(row=r, column=3, value=street_basis).font = BODY_FONT
    for c in range(1, 4):
        ws.cell(row=r, column=c).border = THIN_BORDER
    r += 2

    ws.cell(row=r, column=1, value="AI-Generated Explanation").font = SECTION_FONT
    for c in range(1, 9):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1
    r = _draw_wrapped_box(ws, r, explanation, chars_per_line=100)
    r += 1

    r = _draw_wrapped_box(ws, r,
        "Not investment advice. See the Cover sheet for the full disclaimer.",
        fill=AMBER_FILL_STYLE, border=GOLD_BORDER, font=DISCLAIMER_FONT, chars_per_line=95)

    for col_letter, width in zip("ABCDEFGH", [24, 26, 32, 14, 14, 14, 14, 14]):
        ws.column_dimensions[col_letter].width = width


# ---------------------------------------------------------------
# DCF Forecast sheet
# ---------------------------------------------------------------
def build_forecast_sheet(ws, ticker, company_name, forecast_df: pd.DataFrame):
    _draw_banner(ws, f"{ticker} \u2014 5-Year Free Cash Flow Forecast", company_name)

    headers = list(forecast_df.columns)
    header_row = 4
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_num, value=header)
    _style_header_row(ws, header_row, len(headers))

    for row_num, (_, row) in enumerate(forecast_df.iterrows(), start=header_row + 1):
        for col_num, value in enumerate(row, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if col_num > 1:
                cell.number_format = "$#,##0,,\" M\""
            cell.font = BODY_FONT
            cell.border = THIN_BORDER

    _autofit_columns(ws)


# ---------------------------------------------------------------
# Sensitivity sheet
# ---------------------------------------------------------------
def build_sensitivity_and_scenarios_sheet(ws, ticker, company_name, scenarios: list,
                                             sensitivity_df: pd.DataFrame, current_price: float):
    """Merged sheet: named scenario comparison (a handful of specific,
    labeled judgment calls) sits above the full WACC/growth sensitivity
    grid (every combination, heatmapped) - they were two separate sheets
    covering overlapping ground (the named scenarios are themselves points
    that exist inside the full grid), consolidated after review found the
    workbook had grown to too many thin, single-purpose sheets."""
    _draw_banner(ws, f"{ticker} \u2014 Sensitivity & Scenarios", company_name)

    r = 4
    ws.cell(row=r, column=1, value=f"Current market price: ${current_price:,.2f}").font = LABEL_FONT
    r += 2

    # --- Section 1: Named Scenarios ---
    ws.cell(row=r, column=1, value="Named Scenarios").font = SECTION_FONT
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1
    ws.cell(row=r, column=1,
            value="Base case uses unadjusted historical median assumptions - the most defensible, "
                  "non-circular estimate. Scenarios below show sensitivity to specific, named judgment calls.")
    ws.cell(row=r, column=1).font = DISCLAIMER_FONT
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.row_dimensions[r].height = 32
    r += 2

    headers = ["Scenario", "Capex Assumption", "WACC", "Terminal Growth", "Implied Share Price"]
    scenario_header_row = r
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=scenario_header_row, column=col_num, value=header)
    _style_header_row(ws, scenario_header_row, len(headers))
    r += 1

    for sc in scenarios:
        ws.cell(row=r, column=1, value=sc["label"]).font = BODY_FONT
        ws.cell(row=r, column=2, value=sc.get("capex_label", "Historical median")).font = BODY_FONT
        ws.cell(row=r, column=3, value=sc["wacc"]).number_format = "0.0%"
        ws.cell(row=r, column=4, value=sc["terminal_growth"]).number_format = "0.0%"
        price_cell = ws.cell(row=r, column=5, value=sc["implied_price"])
        price_cell.number_format = "$#,##0.00"
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
        r += 1
    r += 2

    # --- Section 2: Full Sensitivity Grid ---
    ws.cell(row=r, column=1, value="WACC / Terminal Growth Sensitivity Grid").font = SECTION_FONT
    for c in range(1, 1 + len(sensitivity_df.columns) + 1):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 2

    grid_header_row = r
    ws.cell(row=grid_header_row, column=1, value="WACC \\ Growth").font = HEADER_FONT
    for col_num, growth in enumerate(sensitivity_df.columns, start=2):
        ws.cell(row=grid_header_row, column=col_num, value=f"{growth:.1%}")
    _style_header_row(ws, grid_header_row, len(sensitivity_df.columns) + 1)

    for row_offset, (wacc, row) in enumerate(sensitivity_df.iterrows(), start=1):
        rr = grid_header_row + row_offset
        ws.cell(row=rr, column=1, value=f"{wacc:.1%}").font = LABEL_FONT
        for col_offset, value in enumerate(row, start=2):
            cell = ws.cell(row=rr, column=col_offset, value=value)
            cell.number_format = "$#,##0.00"
            cell.font = BODY_FONT
            cell.border = THIN_BORDER

    last_col_letter = get_column_letter(1 + len(sensitivity_df.columns))
    data_range = f"B{grid_header_row + 1}:{last_col_letter}{grid_header_row + len(sensitivity_df)}"
    rule = ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B",
    )
    ws.conditional_formatting.add(data_range, rule)

    _autofit_columns(ws)


# ---------------------------------------------------------------
# Analyst Insights sheet
# ---------------------------------------------------------------
def build_analyst_sheet(ws, ticker, company_name, analyst_data: dict):
    _draw_banner(ws, f"{ticker} \u2014 Analyst Insights", company_name)

    r = 4
    ws.cell(row=r, column=1, value="Analyst Price Targets").font = SECTION_FONT
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1

    targets = analyst_data.get("price_targets")
    if targets:
        labels = [("Low", "low"), ("Mean", "mean"), ("Current", "current"), ("High", "high")]
        for label, key in labels:
            value = targets.get(key)
            ws.cell(row=r, column=1, value=label).font = LABEL_FONT
            if value is not None:
                vcell = ws.cell(row=r, column=2, value=value)
                vcell.number_format = "$#,##0.00"
                vcell.font = BODY_FONT
            r += 1
    else:
        ws.cell(row=r, column=1, value="Not available for this ticker").font = DISCLAIMER_FONT
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Analyst Recommendations (most recent period)").font = SECTION_FONT
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1

    recs = analyst_data.get("recommendations")
    if recs is not None and len(recs) > 0:
        latest = recs.iloc[0]
        headers = ["Strong Buy", "Buy", "Hold", "Sell", "Strong Sell"]
        cols = ["strongBuy", "buy", "hold", "sell", "strongSell"]
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=r, column=col_num, value=header)
        _style_header_row(ws, r, len(headers))
        r += 1
        for col_num, key in enumerate(cols, start=1):
            cell = ws.cell(row=r, column=col_num, value=int(latest.get(key, 0)))
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
        r += 2
    else:
        ws.cell(row=r, column=1, value="Not available for this ticker").font = DISCLAIMER_FONT
        r += 2

    ws.cell(row=r, column=1, value="Latest Rating Actions").font = SECTION_FONT
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1

    ratings = analyst_data.get("latest_ratings")
    if ratings is not None and len(ratings) > 0:
        headers = ["Date", "Firm", "From", "To", "Action"]
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=r, column=col_num, value=header)
        _style_header_row(ws, r, len(headers))
        r += 1
        for idx, row in ratings.iterrows():
            ws.cell(row=r, column=1, value=str(idx.date()) if hasattr(idx, "date") else str(idx)).font = BODY_FONT
            ws.cell(row=r, column=2, value=row.get("Firm", "")).font = BODY_FONT
            ws.cell(row=r, column=3, value=row.get("FromGrade", "")).font = BODY_FONT
            ws.cell(row=r, column=4, value=row.get("ToGrade", "")).font = BODY_FONT
            ws.cell(row=r, column=5, value=row.get("Action", "")).font = BODY_FONT
            for c in range(1, 6):
                ws.cell(row=r, column=c).border = THIN_BORDER
            r += 1
    else:
        ws.cell(row=r, column=1, value="Not available for this ticker").font = DISCLAIMER_FONT
        r += 1
    r += 1

    r = _draw_wrapped_box(ws, r,
        "Sourced from Yahoo Finance analyst coverage data via yfinance. Reflects third-party "
        "sell-side analyst opinions, not this tool's own DCF output, and not investment advice.",
        fill=AMBER_FILL_STYLE, border=GOLD_BORDER, font=DISCLAIMER_FONT, chars_per_line=95)

    _autofit_columns(ws)


# ---------------------------------------------------------------
# Live DCF sheet - genuine editable Excel formulas, not Python values
# ---------------------------------------------------------------
INPUT_FONT = Font(name=FONT_NAME, color="1D4ED8", bold=True, size=10.5)  # blue = input, per standard modeling convention
INPUT_FILL = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
FORMULA_FONT = Font(name=FONT_NAME, color=TEXT_DARK, size=10.5)


def build_live_dcf_sheet(ws, ticker, company_name, assumptions, cash, total_debt,
                           shares_outstanding, current_price, wacc, terminal_growth,
                           tax_rate=0.21, capm_info=None):
    """A genuinely interactive sheet: blue cells are inputs you can type over
    directly in Excel; everything else is a live formula that recalculates
    automatically. Pre-filled with this run's base-case values, but every
    number here is yours to change - unlike the rest of the workbook (which
    holds the audited, Python-computed, tested numbers), this sheet is a
    sandbox for testing your own assumptions."""
    _draw_banner(ws, f"{ticker} \u2014 Live DCF (Editable)", company_name)

    ws.cell(row=3, column=1,
            value="Edit any blue cell below and every downstream number recalculates automatically "
                  "in Excel. This sheet is a sandbox - it starts pre-filled with the same base-case "
                  "assumptions as the Summary sheet, but nothing here is locked.")
    ws.cell(row=3, column=1).font = DISCLAIMER_FONT
    ws.merge_cells("A3:H3")

    def input_cell(row, col, value, fmt=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = INPUT_FONT
        cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        if fmt:
            cell.number_format = fmt
        return cell

    def formula_cell(row, col, formula, fmt=None):
        cell = ws.cell(row=row, column=col, value=formula)
        cell.font = FORMULA_FONT
        cell.border = THIN_BORDER
        if fmt:
            cell.number_format = fmt
        return cell

    # --- Assumptions block (all inputs) ---
    r = 5
    ws.cell(row=r, column=1, value="Key Assumptions (edit the blue cells)").font = SECTION_FONT
    for c in range(1, 4):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1

    input_rows = {
        "revenue_growth": (r, "Revenue Growth", assumptions["avg_revenue_growth"], "0.0%"),
        "ebit_margin": (r + 1, "EBIT Margin", assumptions["avg_ebit_margin"], "0.0%"),
        "capex_pct": (r + 2, "Capex % of Revenue", assumptions["avg_capex_pct_revenue"], "0.0%"),
        "da_pct": (r + 3, "D&A % of Revenue", assumptions["avg_da_pct_revenue"], "0.0%"),
        "nwc_pct": (r + 4, "NWC % of Revenue", assumptions.get("avg_nwc_pct_revenue", 0), "0.0%"),
        "tax_rate": (r + 5, "Tax Rate", tax_rate, "0.0%"),
        "wacc": (r + 6, "WACC (Discount Rate)", wacc, "0.0%"),
        "terminal_growth": (r + 7, "Terminal Growth Rate", terminal_growth, "0.0%"),
    }
    cell_refs = {}
    for key, (row_num, label, value, fmt) in input_rows.items():
        ws.cell(row=row_num, column=1, value=label).font = LABEL_FONT
        input_cell(row_num, 2, value, fmt)
        cell_refs[key] = f"$B${row_num}"

    if capm_info:
        wacc_row = input_rows["wacc"][0]
        note_cell = ws.cell(row=wacc_row, column=4,
                             value=f"Suggested (CAPM): {capm_info['wacc']:.1%}  \u2014  see Summary sheet for the full breakdown")
        note_cell.font = DISCLAIMER_FONT
        ws.merge_cells(start_row=wacc_row, start_column=4, end_row=wacc_row, end_column=8)

    r = input_rows["terminal_growth"][0] + 2

    base_rows = {
        "latest_revenue": (r, "Latest Actual Revenue", assumptions["latest_revenue"], "$#,##0,,\" M\""),
        "latest_nwc": (r + 1, "Latest Actual NWC", assumptions.get("latest_nwc", 0), "$#,##0,,\" M\""),
        "cash": (r + 2, "Cash", cash, "$#,##0,,\" M\""),
        "total_debt": (r + 3, "Total Debt", total_debt, "$#,##0,,\" M\""),
        "shares": (r + 4, "Shares Outstanding", shares_outstanding, "#,##0,,\" M\""),
        "market_price": (r + 5, "Current Market Price (for comparison)", current_price, "$#,##0.00"),
    }
    for key, (row_num, label, value, fmt) in base_rows.items():
        ws.cell(row=row_num, column=1, value=label).font = LABEL_FONT
        input_cell(row_num, 2, value, fmt)
        cell_refs[key] = f"$B${row_num}"
    r = base_rows["market_price"][0] + 2

    # --- 5-year forecast, as live formulas ---
    ws.cell(row=r, column=1, value="5-Year Forecast (all formulas)").font = SECTION_FONT
    for c in range(1, 7):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1
    header_row = r
    ws.cell(row=header_row, column=1, value="Metric")
    for i in range(5):
        ws.cell(row=header_row, column=2 + i, value=f"Year {i + 1}")
    _style_header_row(ws, header_row, 6)

    rev_row = header_row + 1
    ebit_row = rev_row + 1
    nopat_row = ebit_row + 1
    da_row = nopat_row + 1
    capex_row = da_row + 1
    nwc_level_row = capex_row + 1
    nwc_change_row = nwc_level_row + 1
    fcf_row = nwc_change_row + 1
    disc_row = fcf_row + 1
    pv_row = disc_row + 1

    ws.cell(row=rev_row, column=1, value="Revenue").font = LABEL_FONT
    ws.cell(row=ebit_row, column=1, value="EBIT").font = LABEL_FONT
    ws.cell(row=nopat_row, column=1, value="NOPAT").font = LABEL_FONT
    ws.cell(row=da_row, column=1, value="D&A").font = LABEL_FONT
    ws.cell(row=capex_row, column=1, value="Capex").font = LABEL_FONT
    ws.cell(row=nwc_level_row, column=1, value="NWC Level").font = LABEL_FONT
    ws.cell(row=nwc_change_row, column=1, value="Change in NWC").font = LABEL_FONT
    ws.cell(row=fcf_row, column=1, value="Free Cash Flow").font = LABEL_FONT
    ws.cell(row=disc_row, column=1, value="Discount Factor").font = LABEL_FONT
    ws.cell(row=pv_row, column=1, value="PV of FCF").font = LABEL_FONT

    for i in range(5):
        col = 2 + i
        col_letter = get_column_letter(col)
        prev_col_letter = get_column_letter(col - 1)

        rev_formula = (f"={cell_refs['latest_revenue']}*(1+{cell_refs['revenue_growth']})" if i == 0
                        else f"={prev_col_letter}{rev_row}*(1+{cell_refs['revenue_growth']})")
        formula_cell(rev_row, col, rev_formula, "$#,##0,,\" M\"")
        formula_cell(ebit_row, col, f"={col_letter}{rev_row}*{cell_refs['ebit_margin']}", "$#,##0,,\" M\"")
        formula_cell(nopat_row, col, f"={col_letter}{ebit_row}*(1-{cell_refs['tax_rate']})", "$#,##0,,\" M\"")
        formula_cell(da_row, col, f"={col_letter}{rev_row}*{cell_refs['da_pct']}", "$#,##0,,\" M\"")
        formula_cell(capex_row, col, f"={col_letter}{rev_row}*{cell_refs['capex_pct']}", "$#,##0,,\" M\"")
        formula_cell(nwc_level_row, col, f"={col_letter}{rev_row}*{cell_refs['nwc_pct']}", "$#,##0,,\" M\"")

        nwc_change_formula = (f"={col_letter}{nwc_level_row}-{cell_refs['latest_nwc']}" if i == 0
                                else f"={col_letter}{nwc_level_row}-{prev_col_letter}{nwc_level_row}")
        formula_cell(nwc_change_row, col, nwc_change_formula, "$#,##0,,\" M\"")

        formula_cell(fcf_row, col,
                     f"={col_letter}{nopat_row}+{col_letter}{da_row}-{col_letter}{capex_row}-{col_letter}{nwc_change_row}",
                     "$#,##0,,\" M\"")
        formula_cell(disc_row, col, f"=1/(1+{cell_refs['wacc']})^{i + 1}", "0.0000")
        formula_cell(pv_row, col, f"={col_letter}{fcf_row}*{col_letter}{disc_row}", "$#,##0,,\" M\"")

    r = pv_row + 2

    # --- Valuation output, as live formulas ---
    ws.cell(row=r, column=1, value="Valuation (all formulas)").font = SECTION_FONT
    for c in range(1, 3):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1

    last_col_letter = get_column_letter(1 + 5)  # column F = year 5

    pv_explicit_row = r
    ws.cell(row=r, column=1, value="PV of Explicit Period FCFs").font = LABEL_FONT
    formula_cell(r, 2, f"=SUM(B{pv_row}:{last_col_letter}{pv_row})", "$#,##0,,\" M\"")
    r += 1

    tv_row = r
    ws.cell(row=r, column=1, value="Terminal Value").font = LABEL_FONT
    formula_cell(r, 2, f"={last_col_letter}{fcf_row}*(1+{cell_refs['terminal_growth']})/({cell_refs['wacc']}-{cell_refs['terminal_growth']})", "$#,##0,,\" M\"")
    r += 1

    pv_tv_row = r
    ws.cell(row=r, column=1, value="PV of Terminal Value").font = LABEL_FONT
    formula_cell(r, 2, f"=B{tv_row}*{last_col_letter}{disc_row}", "$#,##0,,\" M\"")
    r += 1

    ev_row = r
    ws.cell(row=r, column=1, value="Enterprise Value").font = LABEL_FONT
    formula_cell(r, 2, f"=B{pv_explicit_row}+B{pv_tv_row}", "$#,##0,,\" M\"")
    r += 1

    eq_row = r
    ws.cell(row=r, column=1, value="Equity Value (EV - Debt + Cash)").font = LABEL_FONT
    formula_cell(r, 2, f"=B{ev_row}-{cell_refs['total_debt']}+{cell_refs['cash']}", "$#,##0,,\" M\"")
    r += 2

    price_row = r
    ws.cell(row=r, column=1, value="Implied Share Price").font = Font(name=FONT_NAME, bold=True, size=13, color=NAVY)
    price_cell = formula_cell(r, 2, f"=B{eq_row}/{cell_refs['shares']}", "$#,##0.00")
    price_cell.font = Font(name=FONT_NAME, bold=True, size=13, color=NAVY)
    r += 1

    ws.cell(row=r, column=1, value="vs. Current Market Price").font = LABEL_FONT
    formula_cell(r, 2, f"=B{price_row}/{cell_refs['market_price']}-1", "+0.0%;-0.0%")

    for col_letter, width in zip("ABCDEFGH", [32, 16, 16, 16, 16, 16, 12, 12]):
        ws.column_dimensions[col_letter].width = width


# ---------------------------------------------------------------
# Comps Valuation sheet + Football Field chart
# ---------------------------------------------------------------
from openpyxl.chart import BarChart, Reference


def build_comps_and_football_field_sheet(ws, ticker, company_name, comps_info, dcf_low, dcf_high,
                                            analyst_info, current_price):
    """Merged sheet: comps peer comparison + implied prices sits above the
    football field chart, which visually depends on the comps numbers just
    above it - these were two separate sheets that read as one continuous
    idea anyway (comps data feeding directly into the comparison chart),
    consolidated after review found the workbook had grown to too many
    thin, single-purpose sheets."""
    _draw_banner(ws, f"{ticker} \u2014 Comps & Football Field", company_name)

    r = 4
    ws.cell(row=r, column=1,
            value="A second, independent valuation method: instead of forecasting cash flows, this "
                  "values the company based on what similar public companies currently trade at. "
                  "Peers are chosen deliberately, not auto-detected.")
    ws.cell(row=r, column=1).font = DISCLAIMER_FONT
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 28
    r += 2

    ws.cell(row=r, column=1, value="Peer Comparison").font = SECTION_FONT
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1

    comps_df = comps_info["comps_table"]
    headers = ["Ticker", "Company", "EV/EBITDA", "P/E", "Market Cap"]
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=r, column=col_num, value=header)
    _style_header_row(ws, r, len(headers))
    r += 1

    for idx, row in comps_df.iterrows():
        is_target = idx == ticker
        font = Font(name=FONT_NAME, bold=True, color=NAVY) if is_target else BODY_FONT
        ws.cell(row=r, column=1, value=idx).font = font
        ws.cell(row=r, column=2, value=row.get("Company")).font = font
        ev_ebitda_cell = ws.cell(row=r, column=3, value=row.get("EV/EBITDA"))
        ev_ebitda_cell.font = font
        if pd.notna(row.get("EV/EBITDA")):
            ev_ebitda_cell.number_format = "0.0x"
        pe_cell = ws.cell(row=r, column=4, value=row.get("P/E"))
        pe_cell.font = font
        if pd.notna(row.get("P/E")):
            pe_cell.number_format = "0.0x"
        mcap_cell = ws.cell(row=r, column=5, value=row.get("Market Cap"))
        mcap_cell.font = font
        if pd.notna(row.get("Market Cap")):
            mcap_cell.number_format = "$#,##0,,\" M\""
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Implied Share Price by Method").font = SECTION_FONT
    for c in range(1, 5):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1
    headers2 = ["Method", "Low", "Median/Base", "High"]
    for col_num, header in enumerate(headers2, start=1):
        ws.cell(row=r, column=col_num, value=header)
    _style_header_row(ws, r, len(headers2))
    r += 1

    price_rows = [
        ("EV/EBITDA Multiple", comps_info["implied_price_ev_ebitda_low"],
         comps_info["implied_price_ev_ebitda"], comps_info["implied_price_ev_ebitda_high"]),
        ("P/E Multiple", comps_info["implied_price_pe_low"],
         comps_info["implied_price_pe"], comps_info["implied_price_pe_high"]),
    ]
    for label, low, mid, high in price_rows:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        for col, val in zip([2, 3, 4], [low, mid, high]):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = BODY_FONT
            if val is not None:
                cell.number_format = "$#,##0.00"
            cell.border = THIN_BORDER
        r += 1
    r += 1

    ws.cell(row=r, column=1,
            value="Limitation: peer multiples reflect current market sentiment (including any "
                  "sector-wide optimism or pessimism), not independently-derived fundamentals - "
                  "unlike the DCF, this method can't distinguish 'the whole sector is overvalued' "
                  "from 'this company deserves its current multiple'.")
    ws.cell(row=r, column=1).font = DISCLAIMER_FONT
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 40
    r += 3

    # --- Section 2: Football Field chart ---
    ws.cell(row=r, column=1, value="Football Field: Valuation Range by Method").font = SECTION_FONT
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1
    ws.cell(row=r, column=1,
            value="Each bar shows the low-to-high implied share price range from one valuation "
                  "method. Football field charts are conventionally built as stacked horizontal bars: "
                  "an invisible base segment (the low value) plus a visible range segment on top.")
    ws.cell(row=r, column=1).font = DISCLAIMER_FONT
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 28
    r += 2

    header_row = r
    ws.cell(row=header_row, column=1, value="Method")
    ws.cell(row=header_row, column=2, value="Low (base, hidden)")
    ws.cell(row=header_row, column=3, value="Range (High - Low)")
    ws.cell(row=header_row, column=4, value="Low")
    ws.cell(row=header_row, column=5, value="High")
    _style_header_row(ws, header_row, 5)

    target_price = current_price if current_price else 0
    methods = [
        ("DCF (Base Case \u00b1 Scenarios)", dcf_low, dcf_high),
        ("Comps: EV/EBITDA", comps_info["implied_price_ev_ebitda_low"], comps_info["implied_price_ev_ebitda_high"]),
        ("Comps: P/E", comps_info["implied_price_pe_low"], comps_info["implied_price_pe_high"]),
    ]
    if analyst_info and analyst_info.get("price_targets"):
        targets = analyst_info["price_targets"]
        if targets.get("low") is not None and targets.get("high") is not None:
            methods.append(("Analyst Target Range", targets["low"], targets["high"]))

    rr = header_row + 1
    first_data_row = rr
    for label, low, high in methods:
        low = low if low is not None else 0
        high = high if high is not None else 0
        ws.cell(row=rr, column=1, value=label).font = LABEL_FONT
        ws.cell(row=rr, column=2, value=low).number_format = "$#,##0"
        ws.cell(row=rr, column=3, value=max(high - low, 0)).number_format = "$#,##0"
        ws.cell(row=rr, column=4, value=low).number_format = "$#,##0"
        ws.cell(row=rr, column=5, value=high).number_format = "$#,##0"
        rr += 1
    last_data_row = rr - 1

    ws.cell(row=rr + 1, column=1, value=f"Current Market Price: ${target_price:,.2f}").font = Font(
        name=FONT_NAME, bold=True, color=NAVY)

    # --- Build the stacked horizontal bar chart ---
    chart = BarChart()
    chart.type = "bar"  # horizontal bars
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = f"{ticker} Valuation Range by Method"
    chart.height = 10
    chart.width = 22

    cats = Reference(ws, min_col=1, min_row=first_data_row, max_row=last_data_row)
    base_data = Reference(ws, min_col=2, min_row=header_row, max_row=last_data_row)
    range_data = Reference(ws, min_col=3, min_row=header_row, max_row=last_data_row)
    chart.add_data(base_data, titles_from_data=True)
    chart.add_data(range_data, titles_from_data=True)
    chart.set_categories(cats)

    # IMPORTANT openpyxl quirk: x_axis is ALWAYS the category axis and
    # y_axis is ALWAYS the value axis, regardless of bar orientation - even
    # though a horizontal bar chart visually displays the category axis
    # running down the left side.
    chart.x_axis.title = None
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.scaling.orientation = "maxMin"
    chart.x_axis.axPos = "l"
    chart.y_axis.axPos = "b"
    chart.title.overlay = False

    base_series = chart.series[0]
    base_series.graphicalProperties.noFill = True
    base_series.graphicalProperties.ln.noFill = True

    range_series = chart.series[1]
    range_series.graphicalProperties.solidFill = "0B2545"  # navy, matches the workbook palette

    chart.legend = None
    ws.add_chart(chart, f"A{rr + 3}")

    for col_letter, width in zip("ABCDEF", [28, 16, 18, 14, 14, 14]):
        ws.column_dimensions[col_letter].width = width


# ---------------------------------------------------------------
# Monte Carlo sheet - distribution of outcomes, not one point estimate
# ---------------------------------------------------------------
def build_model_validation_sheet(ws, ticker, company_name, mc_info=None, backtest_info=None, current_price=None):
    """Merged sheet: Monte Carlo (does the model's OUTPUT have a wide range
    of plausible outcomes?) sits above Backtest (does the model's INPUT
    methodology actually predict what happens?) - both are 'how much
    should you trust this' questions, not new valuation numbers, so they
    read naturally as one 'Model Validation' sheet rather than two thin
    single-purpose ones. Either section can be omitted independently -
    this sheet works whether you have Monte Carlo, Backtest, or both."""
    _draw_banner(ws, f"{ticker} \u2014 Model Validation", company_name)

    r = 4

    if mc_info is not None:
        ws.cell(row=r, column=1,
                value=f"{mc_info['n_simulations_completed']:,} simulations, randomizing revenue growth "
                      f"(historical std dev: {mc_info['growth_std_used']:.1%}), WACC (std dev: "
                      f"{mc_info['wacc_std_used']:.1%}), and terminal growth (std dev: "
                      f"{mc_info['terminal_growth_std_used']:.1%}). EBIT margin, capex %, D&A %, and NWC % "
                      f"are held at their historical-median values - a documented v1 simplification.")
        ws.cell(row=r, column=1).font = DISCLAIMER_FONT
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 40
        r += 2

        ws.cell(row=r, column=1, value="Monte Carlo: Distribution Summary").font = SECTION_FONT
        for c in range(1, 3):
            ws.cell(row=r, column=c).fill = SECTION_FILL
        r += 1

        stats_rows = [
            ("Mean Implied Price", mc_info["mean"]),
            ("Median Implied Price", mc_info["median"]),
            ("Std Deviation", mc_info["std"]),
            ("5th Percentile", mc_info["p5"]),
            ("25th Percentile", mc_info["p25"]),
            ("75th Percentile", mc_info["p75"]),
            ("95th Percentile", mc_info["p95"]),
            ("Min", mc_info["min"]),
            ("Max", mc_info["max"]),
        ]
        for label, value in stats_rows:
            ws.cell(row=r, column=1, value=label).font = LABEL_FONT
            vcell = ws.cell(row=r, column=2, value=value)
            vcell.number_format = "$#,##0.00"
            vcell.font = BODY_FONT
            r += 1
        r += 1

        if current_price:
            prob = float((mc_info["prices"] > current_price).mean())
            ws.cell(row=r, column=1,
                    value=f"Probability simulated DCF exceeds current market price (${current_price:,.2f})").font = LABEL_FONT
            pcell = ws.cell(row=r, column=2, value=prob)
            pcell.number_format = "0.0%"
            pcell.font = Font(name=FONT_NAME, bold=True, color=NAVY)
            r += 2

        # --- Histogram: bin the simulated prices, chart the frequency ---
        ws.cell(row=r, column=1, value="Distribution of Outcomes").font = SECTION_FONT
        for c in range(1, 3):
            ws.cell(row=r, column=c).fill = SECTION_FILL
        r += 1

        counts, bin_edges = np.histogram(mc_info["prices"], bins=20)
        header_row = r
        ws.cell(row=header_row, column=1, value="Price Range")
        ws.cell(row=header_row, column=2, value="Frequency")
        _style_header_row(ws, header_row, 2)
        r += 1
        first_data_row = r
        for i in range(len(counts)):
            label = f"${bin_edges[i]:,.0f}-{bin_edges[i+1]:,.0f}"
            ws.cell(row=r, column=1, value=label).font = BODY_FONT
            ws.cell(row=r, column=2, value=int(counts[i])).font = BODY_FONT
            r += 1
        last_data_row = r - 1

        chart = BarChart()
        chart.type = "col"
        chart.title = f"{ticker} Simulated Implied Price Distribution"
        chart.y_axis.title = "Frequency"
        chart.x_axis.title = None
        chart.x_axis.delete = False
        chart.height = 9
        chart.width = 22
        chart.legend = None
        data = Reference(ws, min_col=2, min_row=header_row, max_row=last_data_row)
        cats = Reference(ws, min_col=1, min_row=first_data_row, max_row=last_data_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.solidFill = "0B2545"
        ws.add_chart(chart, f"D{header_row}")

        # Histogram chart is ~9cm tall / occupies rows down to roughly
        # header_row + 20 given 20 bins at default row height - start
        # whatever comes next comfortably below where the chart visually ends.
        r = max(last_data_row, header_row + 22) + 3

    if backtest_info is not None:
        ws.cell(row=r, column=1, value="Backtest: Does the Assumption Methodology Actually Work?").font = SECTION_FONT
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = SECTION_FILL
        r += 1
        ws.cell(row=r, column=1,
                value="Tests the core idea this whole model relies on: does historical-median growth/margin "
                      "predict what actually happens next? Each row predicts a year using ONLY the years "
                      "before it (no lookahead bias), then compares to what actually occurred. Limited to "
                      "the ~4 years of history yfinance provides - directionally informative, not "
                      "statistically robust with this few data points.")
        ws.cell(row=r, column=1).font = DISCLAIMER_FONT
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 42
        r += 2

        ws.cell(row=r, column=1, value="EBIT Margin Predictions").font = SECTION_FONT
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = SECTION_FILL
        r += 1
        headers = ["Year", "Predicted Margin", "Actual Margin", "Error"]
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=r, column=col_num, value=header)
        _style_header_row(ws, r, len(headers))
        r += 1
        for _, row in backtest_info["margin_results"].iterrows():
            ws.cell(row=r, column=1, value=str(row["Year"])).font = BODY_FONT
            for col, key in zip([2, 3, 4], ["Predicted Margin", "Actual Margin", "Error"]):
                cell = ws.cell(row=r, column=col, value=row[key])
                cell.number_format = "+0.00%;-0.00%" if key == "Error" else "0.00%"
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
            r += 1
        if backtest_info["margin_mae"] is not None:
            r += 1
            ws.cell(row=r, column=1, value="Mean Absolute Error").font = LABEL_FONT
            ws.cell(row=r, column=2, value=backtest_info["margin_mae"]).number_format = "0.00%"
            r += 1
            ws.cell(row=r, column=1, value="Bias (negative = under-predicts)").font = LABEL_FONT
            bias_cell = ws.cell(row=r, column=2, value=backtest_info["margin_bias"])
            bias_cell.number_format = "+0.00%;-0.00%"
            bias_cell.font = Font(name=FONT_NAME, bold=True, color=NAVY)
        r += 2

        ws.cell(row=r, column=1, value="Revenue Growth Predictions").font = SECTION_FONT
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = SECTION_FILL
        r += 1
        headers2 = ["Year", "Predicted Growth", "Actual Growth", "Error"]
        for col_num, header in enumerate(headers2, start=1):
            ws.cell(row=r, column=col_num, value=header)
        _style_header_row(ws, r, len(headers2))
        r += 1
        for _, row in backtest_info["growth_results"].iterrows():
            ws.cell(row=r, column=1, value=str(row["Year"])).font = BODY_FONT
            for col, key in zip([2, 3, 4], ["Predicted Growth", "Actual Growth", "Error"]):
                cell = ws.cell(row=r, column=col, value=row[key])
                cell.number_format = "+0.00%;-0.00%" if key == "Error" else "0.00%"
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
            r += 1
        if backtest_info["growth_mae"] is not None:
            r += 1
            ws.cell(row=r, column=1, value="Mean Absolute Error").font = LABEL_FONT
            ws.cell(row=r, column=2, value=backtest_info["growth_mae"]).number_format = "0.00%"
            r += 1
            ws.cell(row=r, column=1, value="Bias (negative = under-predicts)").font = LABEL_FONT
            bias_cell = ws.cell(row=r, column=2, value=backtest_info["growth_bias"])
            bias_cell.number_format = "+0.00%;-0.00%"
            bias_cell.font = Font(name=FONT_NAME, bold=True, color=NAVY)
        r += 2

        r = _draw_wrapped_box(ws, r,
            f"Sample size: {backtest_info['n_margin_tests']} margin predictions, "
            f"{backtest_info['n_growth_tests']} growth predictions. This is a small sample, limited by "
            f"available financial history - treat these results as suggestive, not statistically proven.",
            fill=AMBER_FILL_STYLE, border=GOLD_BORDER, font=DISCLAIMER_FONT, chars_per_line=95)

    for col_letter, width in zip("ABCDEFGH", [26, 18, 16, 14, 14, 14, 14, 14]):
        ws.column_dimensions[col_letter].width = width


# ---------------------------------------------------------------
# Management Assumptions sheet - Base/Upside/Downside scenario planning
# ---------------------------------------------------------------
def build_management_assumptions_sheet(ws, ticker, company_name, scenario_results, current_price):
    """Reframes the DCF's core philosophy: historical data is a STARTING
    POINT for judgment (standard corporate finance / FP&A practice), not
    automatically the forecast itself. Upside/Downside are the company's
    own real best/worst historical years - not fabricated percentages."""
    _draw_banner(ws, f"{ticker} \u2014 Management Assumptions", company_name)

    r = 4
    r = _draw_wrapped_box(ws, r,
        "Historical data is a starting point, not the forecast itself. Base case uses the "
        "historical median (same as the rest of this workbook). Upside and Downside are NOT "
        "arbitrary percentages - they are this company's own actual best and worst historical "
        "years on each metric, so every number in every scenario is something that really happened.",
        chars_per_line=100)
    r += 1

    ws.cell(row=r, column=1, value="Historical Range").font = SECTION_FONT
    for c in range(1, 5):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1
    headers = ["Metric", "Worst Year (Downside)", "Median (Base)", "Best Year (Upside)"]
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=r, column=col_num, value=header)
    _style_header_row(ws, r, len(headers))
    r += 1

    hist = scenario_results["historical_range"]
    range_rows = [
        ("Revenue Growth", hist["growth_min"], hist["growth_median"], hist["growth_max"]),
        ("EBIT Margin", hist["margin_min"], hist["margin_median"], hist["margin_max"]),
        ("Capex % of Revenue", hist["capex_pct_max"], hist["capex_pct_median"], hist["capex_pct_min"]),
    ]
    for label, downside_val, base_val, upside_val in range_rows:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        for col, val in zip([2, 3, 4], [downside_val, base_val, upside_val]):
            cell = ws.cell(row=r, column=col, value=val)
            cell.number_format = "0.0%"
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Implied Share Price by Scenario").font = SECTION_FONT
    for c in range(1, 4):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1
    headers2 = ["Scenario", "Implied Price", "vs. Current Price"]
    for col_num, header in enumerate(headers2, start=1):
        ws.cell(row=r, column=col_num, value=header)
    _style_header_row(ws, r, len(headers2))
    r += 1

    scenario_colors = {"downside": "B91C1C", "base": "1F2937", "upside": "2E7D32"}
    for label in ["downside", "base", "upside"]:
        price = scenario_results[label]["implied_price"]
        gap = price / current_price - 1 if current_price else None
        ws.cell(row=r, column=1, value=label.capitalize()).font = Font(
            name=FONT_NAME, bold=True, color=scenario_colors[label])
        pcell = ws.cell(row=r, column=2, value=price)
        pcell.number_format = "$#,##0.00"
        pcell.font = BODY_FONT
        if gap is not None:
            gcell = ws.cell(row=r, column=3, value=gap)
            gcell.number_format = "+0.0%;-0.0%"
            gcell.font = BODY_FONT
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1
    r += 1

    ws.cell(row=r, column=1, value=f"Current Market Price: ${current_price:,.2f}").font = LABEL_FONT
    r += 2

    r = _draw_wrapped_box(ws, r,
        "Same WACC and terminal growth are used across all three scenarios, isolating the effect "
        "of the OPERATING assumptions (growth, margin, capex) rather than conflating it with a "
        "different discount rate per scenario. See the Live DCF sheet to test your own WACC "
        "alongside these operating scenarios.",
        fill=AMBER_FILL_STYLE, border=GOLD_BORDER, font=DISCLAIMER_FONT, chars_per_line=95)

    for col_letter, width in zip("ABCD", [24, 22, 16, 16]):
        ws.column_dimensions[col_letter].width = width


def export_to_excel(ticker, company_name, current_price, result, assumptions, forecast_df,
                     sensitivity_df, cash, total_debt, shares_outstanding,
                     scenarios=None, ai_output=None, gap_pct=None, analyst_info=None,
                     wacc=0.09, terminal_growth=0.025, tax_rate=0.21, capm_info=None,
                     comps_info=None, monte_carlo_info=None, backtest_info=None,
                     scenario_results=None,
                     output_path="dcf_output.xlsx"):
    wb = Workbook()

    cover_ws = wb.active
    cover_ws.title = "Cover"
    build_cover_sheet(cover_ws, ticker, company_name)

    summary_ws = wb.create_sheet("Summary")
    build_summary_sheet(summary_ws, ticker, company_name, current_price, result,
                         assumptions, cash, total_debt, shares_outstanding, capm_info=capm_info)

    live_ws = wb.create_sheet("Live DCF")
    build_live_dcf_sheet(live_ws, ticker, company_name, assumptions, cash, total_debt,
                          shares_outstanding, current_price, wacc, terminal_growth, tax_rate,
                          capm_info=capm_info)

    if scenario_results is not None:
        mgmt_ws = wb.create_sheet("Management Assumptions")
        build_management_assumptions_sheet(mgmt_ws, ticker, company_name, scenario_results, current_price)

    if comps_info is not None:
        if scenarios:
            scenario_prices = [s["implied_price"] for s in scenarios]
            dcf_low, dcf_high = min(scenario_prices), max(scenario_prices)
        else:
            dcf_low = dcf_high = result["implied_share_price"]

        comps_ff_ws = wb.create_sheet("Comps & Football Field")
        build_comps_and_football_field_sheet(comps_ff_ws, ticker, company_name, comps_info,
                                                dcf_low, dcf_high, analyst_info, current_price)

    if monte_carlo_info is not None or backtest_info is not None:
        validation_ws = wb.create_sheet("Model Validation")
        build_model_validation_sheet(validation_ws, ticker, company_name,
                                       mc_info=monte_carlo_info, backtest_info=backtest_info,
                                       current_price=current_price)

    if analyst_info is not None:
        analyst_ws = wb.create_sheet("Analyst Insights")
        build_analyst_sheet(analyst_ws, ticker, company_name, analyst_info)

    if ai_output is not None:
        ai_ws = wb.create_sheet("AI Valuation Summary")
        build_ai_summary_sheet(ai_ws, ticker, company_name, result, current_price,
                                 gap_pct, ai_output["rating"], ai_output["explanation"],
                                 analyst_info=analyst_info)

    forecast_ws = wb.create_sheet("DCF Forecast")
    build_forecast_sheet(forecast_ws, ticker, company_name, forecast_df)

    if scenarios:
        sens_scen_ws = wb.create_sheet("Sensitivity & Scenarios")
        build_sensitivity_and_scenarios_sheet(sens_scen_ws, ticker, company_name, scenarios,
                                                 sensitivity_df, current_price)
    else:
        # No named scenarios provided - just show the sensitivity grid,
        # reusing the merged function with an empty scenario list rather
        # than maintaining a second standalone code path.
        sens_ws = wb.create_sheet("Sensitivity & Scenarios")
        build_sensitivity_and_scenarios_sheet(sens_ws, ticker, company_name, [],
                                                 sensitivity_df, current_price)

    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    import argparse
    import os
    import yfinance as yf
    from step3_dcf_engine import get_dcf_inputs, get_historical_assumptions, forecast_free_cash_flow
    from step4_sensitivity import build_sensitivity_table
    from config import DEFAULT_WACC, DEFAULT_TERMINAL_GROWTH

    parser = argparse.ArgumentParser(description="Export a full DCF analysis to Excel.")
    parser.add_argument("--ticker", default="MSFT", help="Stock ticker, e.g. MSFT")
    parser.add_argument("--wacc", type=float, default=DEFAULT_WACC, help="Base case discount rate")
    parser.add_argument("--terminal-growth", type=float, default=DEFAULT_TERMINAL_GROWTH, dest="terminal_growth",
                         help="Base case terminal growth rate")
    parser.add_argument("--capex-override", type=float, default=None, dest="capex_override",
                         help="Optional: manually set capex as %% of revenue for an alternate scenario.")
    parser.add_argument("--with-ai", action="store_true", dest="with_ai",
                         help="Include the AI Valuation Summary sheet (calls the Anthropic API - "
                              "requires ANTHROPIC_API_KEY to be set, and incurs a small API cost).")
    parser.add_argument("--peers", default=None,
                         help="Comma-separated list of peer tickers for comps valuation, "
                              "e.g. --peers GOOGL,ORCL,CRM. Adds Comps Valuation and Football "
                              "Field sheets. Free, no API key needed.")
    parser.add_argument("--monte-carlo", type=int, default=None, dest="monte_carlo",
                         help="Number of Monte Carlo simulations to run (e.g. 1000). Adds a "
                              "Monte Carlo sheet showing the distribution of outcomes instead "
                              "of one point estimate. Free, no API key needed.")
    parser.add_argument("--backtest", action="store_true",
                         help="Add a Backtest sheet testing whether historical-median growth/margin "
                              "actually predicted what happened next (leave-future-out validation). "
                              "Free, no API key needed.")
    parser.add_argument("--scenarios", action="store_true",
                         help="Add a Management Assumptions sheet with Base/Upside/Downside cases, "
                              "each derived from this company's own real historical best/worst years "
                              "(not fabricated). Free, no API key needed.")
    args = parser.parse_args()

    ticker = args.ticker
    df = get_dcf_inputs(ticker)

    from step3_dcf_engine import run_dcf_scenario
    from config import DEFAULT_SENSITIVITY_WACC_RANGE, DEFAULT_SENSITIVITY_GROWTH_RANGE

    # get_historical_assumptions/forecast_free_cash_flow don't need cash,
    # debt, or shares - only the final calculate_dcf_valuation step does.
    # So these two run first to determine latest_year, which is needed to
    # look up cash/debt from the DataFrame, before the full scenario runs.
    assumptions = get_historical_assumptions(df)
    forecast_df = forecast_free_cash_flow(assumptions)

    latest_year = assumptions["latest_year"]
    cash = df.loc["Cash", latest_year]
    total_debt = df.loc["Total Debt", latest_year]
    stock_info = yf.Ticker(ticker).info
    shares_outstanding = stock_info.get("sharesOutstanding")
    current_price = stock_info.get("currentPrice")
    company_name = stock_info.get("longName", ticker)

    base = run_dcf_scenario(df, cash=cash, total_debt=total_debt,
                              shares_outstanding=shares_outstanding,
                              wacc=args.wacc, terminal_growth=args.terminal_growth)
    result = base["result"]
    gap_pct = result["implied_share_price"] / current_price - 1

    sensitivity_df = build_sensitivity_table(
        forecast_df, cash, total_debt, shares_outstanding,
        wacc_range=DEFAULT_SENSITIVITY_WACC_RANGE,
        growth_range=DEFAULT_SENSITIVITY_GROWTH_RANGE,
    )

    def build_scenario_entry(label, wacc, terminal_growth, capex_override=None):
        overrides = {"avg_capex_pct_revenue": capex_override} if capex_override is not None else None
        scenario = run_dcf_scenario(df, cash=cash, total_debt=total_debt,
                                      shares_outstanding=shares_outstanding,
                                      wacc=wacc, terminal_growth=terminal_growth,
                                      overrides=overrides)
        return {
            "label": label,
            "capex_label": f"{capex_override:.0%}" if capex_override is not None else "Historical median",
            "wacc": wacc,
            "terminal_growth": terminal_growth,
            "implied_price": scenario["result"]["implied_share_price"],
        }

    scenarios = [
        build_scenario_entry("Base case", args.wacc, args.terminal_growth),
        build_scenario_entry(f"Lower WACC ({args.wacc - 0.02:.1%})", args.wacc - 0.02, args.terminal_growth),
        build_scenario_entry(f"Higher WACC ({args.wacc + 0.02:.1%})", args.wacc + 0.02, args.terminal_growth),
    ]
    if args.capex_override is not None:
        scenarios.append(build_scenario_entry(
            f"Custom capex override ({args.capex_override:.0%} of revenue)",
            args.wacc, args.terminal_growth, capex_override=args.capex_override,
        ))

    from analyst_data import get_analyst_data
    print("Pulling analyst insights (free, no API key needed)...")
    analyst_info = get_analyst_data(ticker)

    from capm_wacc import compute_capm_wacc
    print("Computing suggested WACC via CAPM (free, no API key needed)...")
    capm_info = compute_capm_wacc(ticker, df, stock_info, tax_rate=0.21)

    ai_output = None
    if args.with_ai:
        if not os.environ.get("ANTHROPIC_API_KEY"):
            print("--with-ai was set but ANTHROPIC_API_KEY is not set. Skipping AI sheet.")
        else:
            from step6_ai_summary import generate_gap_explanation
            print("Generating AI valuation summary (this calls the Anthropic API)...")
            ai_output = generate_gap_explanation(
                ticker, df, assumptions, result, current_price, stock_info,
                args.wacc, args.terminal_growth, analyst_info=analyst_info,
            )

    comps_info = None
    if args.peers:
        from comps_valuation import compute_comps_valuation
        peer_tickers = [p.strip().upper() for p in args.peers.split(",")]
        print(f"Pulling comps data for peers: {peer_tickers} (free, no API key needed)...")
        comps_info = compute_comps_valuation(
            ticker, peer_tickers, shares_outstanding, cash, total_debt,
        )

    monte_carlo_info = None
    if args.monte_carlo:
        from monte_carlo import run_monte_carlo
        print(f"Running {args.monte_carlo} Monte Carlo simulations (free, no API key needed)...")
        monte_carlo_info = run_monte_carlo(
            df, cash=cash, total_debt=total_debt, shares_outstanding=shares_outstanding,
            wacc_base=args.wacc, terminal_growth_base=args.terminal_growth,
            n_simulations=args.monte_carlo,
        )
        print(f"  Completed {monte_carlo_info['n_simulations_completed']} / {monte_carlo_info['n_simulations_requested']}")

    backtest_info = None
    if args.backtest:
        from backtest import run_backtest
        print("Running backtest (free, no API key needed)...")
        backtest_info = run_backtest(df)

    mgmt_scenario_results = None
    if args.scenarios:
        from scenario_planning import run_scenario_valuations
        print("Building Base/Upside/Downside scenarios from real historical data (free, no API key needed)...")
        mgmt_scenario_results = run_scenario_valuations(
            df, cash=cash, total_debt=total_debt, shares_outstanding=shares_outstanding,
            wacc=args.wacc, terminal_growth=args.terminal_growth,
        )

    export_to_excel(ticker, company_name, current_price, result, assumptions, forecast_df,
                     sensitivity_df, cash, total_debt, shares_outstanding,
                     scenarios=scenarios, ai_output=ai_output, gap_pct=gap_pct,
                     analyst_info=analyst_info, capm_info=capm_info, comps_info=comps_info,
                     monte_carlo_info=monte_carlo_info, backtest_info=backtest_info,
                     scenario_results=mgmt_scenario_results,
                     wacc=args.wacc, terminal_growth=args.terminal_growth,
                     output_path=f"{ticker}_dcf_output.xlsx")
